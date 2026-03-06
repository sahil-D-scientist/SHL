"""
Evaluation pipeline for the SHL Assessment Recommendation Engine.
Computes Recall@K on the train set and generates predictions for the test set.

Uses threading to parallelize query_analyzer + retriever across all queries,
then runs reranker sequentially.

All outputs are saved to the output/ folder.
"""

import csv
import os
import sys
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

import openpyxl

import config
from core.graph import (
    recommend,
    query_analyzer_node,
    retriever_node,
    reranker_node,
    GraphState,
    get_bm25,
    warmup,
)


class TeeWriter:
    """Write to both a file and stdout simultaneously."""
    def __init__(self, file_path):
        self.file = open(file_path, "w", encoding="utf-8")
        self.stdout = sys.stdout

    def write(self, text):
        self.stdout.write(text)
        self.file.write(text)

    def flush(self):
        self.stdout.flush()
        self.file.flush()

    def close(self):
        self.file.close()


def normalize_url(url: str) -> str:
    """Normalize URL for comparison (remove trailing slash, standardize domain)."""
    url = url.strip().rstrip("/").lower()
    url = url.replace("/solutions/products/product-catalog/view/",
                      "/products/product-catalog/view/")
    url = url.replace("www.shl.com/products/", "www.shl.com/products/")
    return url


def load_train_set(filepath: str) -> dict[str, list[str]]:
    """Load train set: {query: [list of relevant URLs]}."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Train-Set"]

    query_urls = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        query = ws.cell(row=row, column=1).value
        url = ws.cell(row=row, column=2).value
        if query and url:
            query_urls[query].append(normalize_url(url))

    return dict(query_urls)


def load_test_set(filepath: str) -> list[str]:
    """Load test set queries."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Test-Set"]

    queries = []
    for row in range(2, ws.max_row + 1):
        query = ws.cell(row=row, column=1).value
        if query:
            queries.append(query)

    return queries


def compute_recall_at_k(recommended_urls: list[str], relevant_urls: list[str], k: int = 10) -> float:
    """Compute Recall@K for a single query."""
    recommended_normalized = [normalize_url(u) for u in recommended_urls[:k]]
    relevant_normalized = [normalize_url(u) for u in relevant_urls]

    if not relevant_normalized:
        return 0.0

    hits = sum(1 for url in relevant_normalized if url in recommended_normalized)
    return hits / len(relevant_normalized)


def _run_retrieval(query: str, idx: int, total: int) -> dict:
    """Run query_analyzer + retriever for a single query (thread-safe)."""
    print(f"  [Retrieval {idx}/{total}] Starting: {query[:60]}...")
    state = GraphState(
        query=query,
        search_queries=[],
        skills=[],
        max_duration=None,
        domain="",
        candidates=[],
        recommendations=[],
    )
    s1 = query_analyzer_node(state)
    state.update(s1)
    s2 = retriever_node(state)
    state.update(s2)
    print(f"  [Retrieval {idx}/{total}] Done: {len(state['candidates'])} candidates")
    return state


def evaluate_train_set(dataset_path: str):
    """Evaluate on the train set with parallel retrieval."""
    K = config.TOP_K_FINAL

    print(f"Loading train set... (evaluating Recall@{K})")
    query_urls = load_train_set(dataset_path)
    print(f"Loaded {len(query_urls)} queries with {sum(len(v) for v in query_urls.values())} total labels")

    print("\nWarming up models...")
    warmup()
    get_bm25()
    print("Models ready.\n")

    queries = list(query_urls.keys())

    # Phase 1: Parallel retrieval (query_analyzer + retriever)
    print("=" * 40)
    print("PHASE 1: Parallel Retrieval")
    print("=" * 40)
    states = {}
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {
            executor.submit(_run_retrieval, q, i + 1, len(queries)): q
            for i, q in enumerate(queries)
        }
        for future in as_completed(futures):
            q = futures[future]
            states[q] = future.result()

    # Phase 2: Sequential reranking (LLM calls)
    print(f"\n{'=' * 40}")
    print("PHASE 2: Sequential Reranking")
    print("=" * 40)

    recalls = []
    for i, query in enumerate(queries, 1):
        relevant_urls = query_urls[query]
        state = states[query]

        print(f"\n--- Query {i}/{len(queries)} ---")
        print(f"Query: {query[:80]}...")
        print(f"Relevant URLs: {len(relevant_urls)}")

        result = reranker_node(state)
        recommendations = result["recommendations"]
        recommended_urls = [r["url"] for r in recommendations]

        recall = compute_recall_at_k(recommended_urls, relevant_urls, k=K)
        recalls.append(recall)

        print(f"Recommended: {len(recommended_urls)} assessments")
        print(f"Recall@{K}: {recall:.4f}")

        rec_normalized = [normalize_url(u) for u in recommended_urls]
        for url in relevant_urls:
            match = "HIT" if normalize_url(url) in rec_normalized else "MISS"
            name = url.split("/view/")[-1].rstrip("/")
            print(f"  [{match}] {name}")

    mean_recall = sum(recalls) / len(recalls) if recalls else 0.0
    print(f"\n{'=' * 60}")
    print(f"Mean Recall@{K}: {mean_recall:.4f}")
    print(f"Per-query recalls: {[f'{r:.2f}' for r in recalls]}")
    return mean_recall


def generate_test_predictions(dataset_path: str, output_path: str):
    """Generate predictions CSV for the test set."""
    print("\nLoading test set...")
    queries = load_test_set(dataset_path)
    print(f"Loaded {len(queries)} test queries")

    rows = []
    for i, query in enumerate(queries, 1):
        print(f"\n--- Test Query {i}/{len(queries)} ---")
        print(f"Query: {query[:80]}...")

        recommendations = recommend(query)
        print(f"Generated {len(recommendations)} recommendations")

        for rec in recommendations:
            rows.append([query, rec["url"]])

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Query", "Assessment_url"])
        writer.writerows(rows)

    print(f"\nPredictions saved to {output_path}")
    print(f"Total rows: {len(rows)}")


if __name__ == "__main__":
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)

    dataset_path = os.path.join(os.path.dirname(__file__), "input", "Gen_AI Dataset.xlsx")
    # Tee all console output to a log file
    log_path = os.path.join(config.OUTPUT_DIR, "evaluation_log.txt")
    tee = TeeWriter(log_path)
    sys.stdout = tee

    print(f"SHL Assessment Recommendation Engine - Evaluation")
    print(f"Model: {config.LLM_MODEL}")
    print(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # --- Train set evaluation ---
    print(f"\n{'=' * 60}")
    print("EVALUATION ON TRAIN SET")
    print("=" * 60)
    mean_recall = evaluate_train_set(dataset_path)

    # # --- Test set predictions ---
    # print(f"\n\n{'=' * 60}")
    # print("GENERATING TEST SET PREDICTIONS")
    # print("=" * 60)
    # predictions_path = os.path.join(config.OUTPUT_DIR, "predictions.csv")
    # generate_test_predictions(dataset_path, predictions_path)

    # # --- Summary ---
    # print(f"\n{'=' * 60}")
    # print("OUTPUT FILES GENERATED:")
    # print(f"  1. {log_path}")
    # print(f"  2. {predictions_path}")
    # print(f"  Mean Recall@{config.TOP_K_FINAL}: {mean_recall:.4f}")
    # print("=" * 60)

    # Restore stdout and close log
    sys.stdout = tee.stdout
    tee.close()

    print(f"\nDone! All outputs saved to: {config.OUTPUT_DIR}")
    print(f"  - evaluation_log.txt  (full evaluation log with train Recall@K)")
    print(f"  - predictions.csv     (test set predictions)")
    print(f"  Mean Recall@{config.TOP_K_FINAL}: {mean_recall:.4f}")
